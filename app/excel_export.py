import io
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.styles.borders import Border, Side, BORDER_THIN


def excel_export(user_inputs, ten_table_data, bar_chart_x_axis_values, bar_chart_y_axis_values):
    wb = Workbook()

    # ---- create styles/formats to be used later ----
    # for applicable numerical values
    comma_no_decimal_style = NamedStyle(name='comma_no_decimal_style')
    comma_no_decimal_style.number_format = '#,##0'
    comma_no_decimal_style.alignment = Alignment(horizontal='left')   # left align like text is by default in Excel

    # for table header
    ten_table_header_style = NamedStyle(name='ten_table_header_style')
    ten_table_header_style.font = Font(bold=True)
    ten_table_header_style.fill = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
    ten_table_header_style.border = Border(top=Side(border_style=BORDER_THIN, color='00000000'), bottom=Side(border_style=BORDER_THIN, color='00000000'),
                                     left=Side(border_style=BORDER_THIN, color='00000000'), right=Side(border_style=BORDER_THIN, color='00000000'))

    # for table title
    ten_table_title_style = NamedStyle(name='ten_table_title_style')
    ten_table_title_style.font = Font(size=18, bold=True)
    ten_table_title_style.alignment = Alignment(horizontal='center')

    # -----------------------INPUT TAB-----------------------
    ws_input = wb.active
    ws_input.title = 'input'

    # provide user selections
    for row, label in enumerate(['Rank Position', 'Rank By', 'State', 'City', 'Zip Code', 'Place of Service', 'Provider Type', 'Credential', 'HCPCS Code'], start=1):
        ws_input.cell(column=1, row=row, value=label).font = Font(bold=True)

    ws_input.cell(column=2, row=1, value=', '.join(user_inputs['rank_position']) if 'rank_position' in user_inputs else '')     # (rank position should always be in user inputs, since it's required)
    ws_input.cell(column=2, row=2, value=', '.join(user_inputs['rank_by']) if 'rank_by' in user_inputs else '')                 # (rank by should always be in user inputs, since it's required)
    ws_input.cell(column=2, row=3, value=', '.join(user_inputs['state']) if 'state' in user_inputs else '(all)')
    ws_input.cell(column=2, row=4, value=', '.join(user_inputs['city']) if 'city' in user_inputs else '(all)')
    ws_input.cell(column=2, row=5, value=', '.join(user_inputs['zip_code']) if 'zip_code' in user_inputs else '(all)')
    ws_input.cell(column=2, row=6, value=', '.join(user_inputs['place_of_service']) if 'place_of_service' in user_inputs else '(all)')
    ws_input.cell(column=2, row=7, value=', '.join(user_inputs['provider_type']) if 'provider_type' in user_inputs else '(all)')
    ws_input.cell(column=2, row=8, value=', '.join(user_inputs['credential']) if 'credential' in user_inputs else '(all)')
    ws_input.cell(column=2, row=9, value=', '.join(user_inputs['hcpcs_code']) if 'hcpcs_code' in user_inputs else '(all)')

    for col in ['A', 'B']:
        ws_input.column_dimensions[col].width = 20

    # -----------------------DATA TAB-----------------------
    ws_data = wb.create_sheet(title='data')

    # ----provide table data----
    for col, label in enumerate(['Provider ID', 'Patients', 'Avg Charged', 'Avg Allowed', 'Avg Paid'], start=1):
        ws_data.cell(column=col, row=1, value=label)

    for row, data in enumerate(ten_table_data, start=2):
        ws_data.cell(column=1, row=row, value=data['provider_id'])
        ws_data.cell(column=2, row=row, value=int(data['patients'].replace(',', ''))).style = comma_no_decimal_style      # transforming string into number (so user can more easily perform calculations in Excel) and formatting
        ws_data.cell(column=3, row=row, value=int(data['avg_charged'].replace(',', ''))).style = comma_no_decimal_style
        ws_data.cell(column=4, row=row, value=int(data['avg_allowed'].replace(',', ''))).style = comma_no_decimal_style
        ws_data.cell(column=5, row=row, value=int(data['avg_paid'].replace(',', ''))).style = comma_no_decimal_style

    # ----provide chart data----
    ws_data.cell(column=8, row=1, value='Bar Chart')

    for row, x in enumerate(bar_chart_x_axis_values, start=2):
        ws_data.cell(column=8, row=row, value=x)

    for row, y in enumerate(bar_chart_y_axis_values, start=2):
        ws_data.cell(column=9, row=row, value=y).style = comma_no_decimal_style

    # set column width and bold headings for table & chart data
    for col in ['A', 'B', 'C', 'D', 'E', 'H', 'I']:
        ws_data.column_dimensions[col].width = 16
        ws_data[f'{col}1'].font = Font(bold=True)

    # -----------------------RESULTS TAB-----------------------
    ws_results = wb.create_sheet(title='results')

    # ---- create and format table ----
    ten_table_title = f"{user_inputs['rank_position'][0]} 10 Providers by {'Number of Patients' if user_inputs['rank_by'][0] == 'Patients' else 'Average Charged Amount'}"
    ws_results.merge_cells('A1:E1')
    ws_results.cell(column=1, row=1, value=ten_table_title).style = ten_table_title_style

    for col, label in enumerate(['Provider ID', 'Patients', 'Avg Charged', 'Avg Allowed', 'Avg Paid'], start=1):
        ws_results.cell(column=col, row=2, value=label).style = ten_table_header_style

    for row in range(3, len(ten_table_data)+3):
        ws_results.cell(column=1, row=row, value=f"=data!A{row-1}")
        ws_results.cell(column=2, row=row, value=f"=data!B{row-1}").style = comma_no_decimal_style
        ws_results.cell(column=3, row=row, value=f"=data!C{row-1}").style = comma_no_decimal_style
        ws_results.cell(column=4, row=row, value=f"=data!D{row-1}").style = comma_no_decimal_style
        ws_results.cell(column=5, row=row, value=f"=data!E{row-1}").style = comma_no_decimal_style

    ten_table = Table(displayName="TenTable", ref=f'A3:E{len(ten_table_data)+2}', headerRowCount=0)     # didn't specify a header so that default column filters aren't created
    ten_table.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
    ws_results.add_table(ten_table)

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_results.column_dimensions[col].width = 16

    # ---- create bar chart ----
    bar_chart = BarChart()
    bar_chart.type = 'col'
    bar_chart.style = 10
    bar_chart.title = 'Total Patients by Average Charged Amount'
    bar_chart.y_axis.title = 'Patients'
    bar_chart.x_axis.title = 'Avg Charged'
    bar_chart.add_data(Reference(ws_data, min_col=9,  max_col=9, min_row=2, max_row=len(bar_chart_y_axis_values)+1), titles_from_data=False)
    bar_chart.set_categories(Reference(ws_data, min_col=8, max_col=8, min_row=2, max_row=len(bar_chart_x_axis_values)+1))
    bar_chart.shape = 4
    bar_chart.legend = None
    bar_chart.height = 9
    bar_chart.width = 18
    ws_results.add_chart(bar_chart, 'H1')

    # ---- add footnotes ----
    footnote1 = "Based on 2017 Medicare Provider Utilization and Payment Data accessed December 18, 2019 from data.cms.gov:  https://data.cms.gov/Medicare-Physician-Supplier/Medicare-Provider-Utilization-and-Payment-Data-Phy/fs4p-t5eq."
    footnote2 = "Only providers listed as individuals and their services within the 50 states/DC are included.  Please note that Provider ID is a generated number created by the developer in an effort to de-identify providers."
    footnote3 = "Additional data cleaning/transformations on the data set were performed as needed at the sole discretion of the developer."
    footnote4 = "The specific data shown in the above table & graph is based on user selections (see input tab) in the web application."

    ws_results.cell(column=1, row=20, value='Source Data:').font = Font(size=9, underline='single')

    for i, note in enumerate([footnote1, footnote2, footnote3], start=1):
        ws_results.cell(column=1, row=20+i, value=note).font = Font(size=9)

    ws_results.cell(column=1, row=25, value=footnote4).font = Font(size=9)      # writes a couple of lines down from the previous footnote in order to leave a blank line between this last footnote and the prior ones

    # set the results sheet (which is at index 2) as the active worksheet so that the user will be on this sheet when opening the file
    wb.active = 2   # the active worksheet is now the one at index 2

    # save workbook to a stream (vs. saving to a file on disk) to later send to user
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # go to the beginning of the stream

    return excel_stream